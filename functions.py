# functions.py
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from playwright.sync_api import sync_playwright
from playwright._impl._errors import Error, TimeoutError, TargetClosedError
import pandas as pd
import time
import os
from _datetime import datetime
from tkinter import filedialog

# GLOBAL VARIABLES (unchanged)
CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
USER_DATA_DIR = r"C:/Users/star1/Desktop/PlaywrightChromeUser"

### 0 - HELPER FUNCTIONS ###
def clean_price_spans(value):
    if not value:
        return  None

    if isinstance(value, (int, float)):
        return float(value)  # Already a number no need for conversion Lukckily :)


    # value = re.sub(r'[^\d,.-]', '', str(value))  # Remove non-numeric characters
    # value = value.replace(',', '.')

    value = str(value)
    if value.endswith("dh"):
        value = value.split("dh")[0].strip()

    value = value.replace('\xa0', '')  # Remove non-breaking space
    value = value.replace(' ', '')  # Remove standard space (1 845)
    value = value.replace(",", ".")  # Replace comma decimal (1.845,35)

    try:
        return float(value)
    except ValueError:
        print(f"⚠️ Unable to convert '{value}' to float.")
        return None

### Updated set_info_msgs to accept an update_ui_callback ###
def set_info_msgs(update_ui_callback, theme, message):
    update_ui_callback(theme, message)
    return True

### INITIATE_WORKFLOW()
def initiate_wf(app_instance):
    app_instance.step_labels[0].configure(text="Selection Fichier")
    app_instance.step_labels[1].configure(text="Extraction Données")
    app_instance.step_labels[2].configure(text="Processus WAT.ERP")
    app_instance.step_labels[3].configure(text="Export Données")
    set_wf_state(app_instance, 1)

def set_wf_state(app_instance, current_step):
        for i in range(4):
            step = i+1
            if  step < current_step:#step is complete
                app_instance.number_frames[i].configure(fg_color='#28a745')
                app_instance.number_labels[i].configure(text_color='#FFF')
                app_instance.step_labels[i].configure(text_color ='#28a745')

            elif step > current_step: #step is inactive
                app_instance.number_frames[i].configure(fg_color='#e9ecef')
                app_instance.number_labels[i].configure(text_color='#333')
                app_instance.step_labels[i].configure(text_color ='#333')
            else: #step is ongoing
                app_instance.number_frames[i].configure(fg_color='#007bff')
                app_instance.number_labels[i].configure(text_color='#FFF')
                app_instance.step_labels[i].configure(text_color ='#007bff')

### 1 - OPEN FILE DIALOG ###

def openfile_dialog(app_instance):
    app_instance.file_textbox.delete(0, 'end')
    file_name = filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx .xls")])

    if not file_name:
        app_instance.file_path_label.configure(text="Fichier :")
        app_instance.file_textbox.insert(0, "No File Selected")
        return -1

    try:
        excel_file = pd.ExcelFile(file_name)
        sheet_names = excel_file.sheet_names
        app_instance.sheet_combo.configure(values=sheet_names, state="readonly")

        if sheet_names:
            app_instance.sheet_combo.set(sheet_names[0])

    except Exception as e:
        app_instance.set_info_msgs("alert", f"Error reading Excel file: {e}")
        app_instance.sheet_combo.configure(values=["_____"])
        return -1

    app_instance.file_path_label.configure(text=f"Fichier : {file_name}")
    app_instance.file_textbox.insert(0, f"{os.path.basename(file_name)}")
    app_instance.set_info_msgs('success', f'Fichier Selectionné {os.path.basename(file_name)}')
    set_wf_state(app_instance,2)
    return 1



### 2 EXTRACT DATA

def extract_data (app_instance): # Renamed 'self' to 'app_instance' for clarity

    file_path_text = app_instance.file_path_label.cget("text")
    clean_path = file_path_text.replace("Fichier : ", "", 1).strip() if file_path_text.startswith("Fichier : ") else file_path_text.strip()
    selected_sheet = app_instance.sheet_combo.get()
    radio_option = app_instance.radio_option.get()

    if radio_option == 'print': columns = ['N° de facture', 'Catégorie de facturation']
    elif radio_option == 'pns': columns = ['N° de contrat']
    else: columns = None
    try: df = pd.read_excel(clean_path, sheet_name=selected_sheet, usecols=columns)

    except FileNotFoundError:
        app_instance.set_info_msgs('alert', 'Fichier Introuvable') # Calls app_instance's method
        return -1

    except ValueError as ve:
        if "Worksheet" in str(ve):
            app_instance.set_info_msgs('alert', 'Feuille Excel invalide') # Calls app_instance's method
            return -2

        if "Usecols" in str(ve):
            app_instance.set_info_msgs('alert', 'Colonnes manquantes') # Calls app_instance's method
            return -3

        app_instance.set_info_msgs('alert', 'Lecture Excel') # Calls app_instance's method
        return -9
    except Exception:
        return -9

    if app_instance.radio_option.get() == 'print':
        df['N° de facture'] = df['N° de facture'].fillna('0').astype(str).str.rstrip('.0')
        df['Catégorie de facturation'] = df['Catégorie de facturation'].fillna('').astype(str)
        data_array = list(zip(df['N° de facture'], df['Catégorie de facturation']))

        app_instance.set_info_msgs('info', 'Données des Factures extraites') # Calls app_instance's method

    else:
        df['N° de contrat'] = df['N° de contrat'].fillna('0').astype(str).str.replace(r'\.0$', '', regex = True)
        data_array = df['N° de contrat'].tolist()

        app_instance.set_info_msgs('info', 'Données des Contrats extraites') # Calls app_instance's method

    if isinstance(data_array, list) and len(data_array) > 0:
        app_instance.execute_button.configure(state = "normal")
        set_wf_state(app_instance,3)
    return data_array

## 3 - PLAYWRIGHT PROCESSING##
### 3 - 1  LOGIN---
def waterp_login(username, password, url="https://waterp-cas.srm-cas.local/"): # No change needed here if it doesn't call set_info_msgs
    login_tag = 0
    storage = None
    info_message = ""

    with sync_playwright() as p:

        browser = p.chromium.launch(headless=False, executable_path=CHROME_PATH)
        context = browser.new_context(ignore_https_errors=True)
        page = context.new_page ()
        try:
            page.goto(url, timeout=10000)
            page.fill("input[name='username']", username)
            page.fill("input[name='password']", password)
            page.click("form[name = 'form'] > button")
            page.wait_for_timeout(2000)

            alert_text = page.locator(".form-container .row .alert-danger")
            if alert_text.is_visible():
                info_message = f"Connexion Échouée : {alert_text.inner_text()} !"
                login_tag = -1
            else:
                page.wait_for_selector("a[href='#clientèleContent']")
                storage = context.storage_state(path="state.json")
                info_message = "Connexion avec Succès !"
                login_tag = 1
            browser.close()

        except Exception as e:
            error_str = str(e)
            storage = None
            login_tag = -2
            if "ERR_INTERNET_DISCONNECTED" in str(error_str) or  "ERR_NAME_NOT_RESOLVED" in str(error_str): info_message = "Erreur de connexion au réseau"
            elif "ERR_ABORTED" in str(error_str) or "ERR_NETWORK_CHANGED" in str(error_str) or "has been closed" in str(error_str): info_message = "Connexion Coupée"
            else: info_message = str(error_str)

        return login_tag,storage, info_message

# Process functions now accept set_info_msgs_callback
def process_factures(set_info_msgs_callback, factures, progress_callback,print_logo=False, advanced = None):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, executable_path=CHROME_PATH)
        context = browser.new_context(storage_state="state.json", ignore_https_errors=True)
        context.set_default_navigation_timeout(60000)
        page = context.new_page()
        page.goto("https://waterp-cas.srm-cas.local/reporting/report")
        page.wait_for_load_state("domcontentloaded")
        page.click('.reportingnav--types > li > a:has-text("Facturation")')
        page.wait_for_load_state("domcontentloaded")

        if advanced == None:
            total_factures = len(factures)
            start_time = None
            end_time = None

            for i, facture in enumerate(factures):
                if i == 0:
                    start_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

                if facture[1].startswith("BT "):
                    page.click("div#reportDoc_FCEP200U_waterp_client_cas_cas > a")
                    page.fill('input[id$="NUM_FAC_CLI"]', facture[0])
                    page.wait_for_load_state("domcontentloaded")
                    if print_logo:
                        page.fill('input[id$="P_Logo"]', 'O')

                    else:
                        page.fill('input[id$="P_Logo"]', '')

                    page.click("button[id ='button-FCEP200U_waterp_client_cas_cas']")
                    page.wait_for_timeout(2000)
                    page.click("div#reportDoc_FCEP200U_waterp_client_cas_cas > a")

                elif facture[1].startswith("MT "):
                    page.click("div#reportDoc_FCEP200U_MT_waterp_client_cas_cas > a")
                    page.fill('input[id$="num_fac_cli"]', facture[0])
                    page.wait_for_load_state("domcontentloaded")
                    if print_logo:

                        page.fill('input[id$="P_Logo"]', 'O')
                    else:
                        page.fill('input[id$="P_Logo"]', '')
                    page.click("button[id ='button-FCEP200U_MT_waterp_client_cas_cas']")
                    page.wait_for_timeout(2000)
                    page.click("div#reportDoc_FCEP200U_MT_waterp_client_cas_cas > a")

                progress_percent = f'{((i + 1) / total_factures) * 100:.0f}%' # Format as integer percentage
                progress_text = f'{i + 1}/{total_factures}'
                progress_callback(progress_text, progress_percent) # Call the progress callback
                time.sleep(0.1) # Reduced sleep for better responsiveness during simulation

                if i == total_factures - 1:
                    end_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            set_info_msgs_callback('info', f'start_time: {start_time} \n end_time: {end_time}')
            set_info_msgs_callback('info', f'{total_factures} proccessed')
            browser.close()
        #else:
            #champ_num_client = advanced['num_client']
            #loop over période and année
    return True, start_time, end_time

def process_contrats(set_info_msgs_callback, contrats, progress_callback):

    fetched_data_list = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, executable_path=CHROME_PATH)
        context = browser.new_context(storage_state="state.json", ignore_https_errors=True)
        context.set_default_navigation_timeout(60000)
        page = context.new_page()

        total_contrats = len(contrats)
        for i, contrat in enumerate(contrats):
            try:
                # page.goto(r'file:///C:/Users/star1/Downloads/Wat.erp%20-%20Synth%C3%A8se%20client+test.mhtml')  # lien irl = synthese/contrat/NUM_CONTRAT

                page.goto(fr'https://waterp-cas.srm-cas.local/clientele/synthese/contrat/{contrat}')
                page.wait_for_load_state('domcontentloaded')
                page.click('a[data-target="#compteclient"]')
                page.wait_for_load_state("load")
                empty_solde = page.locator('#compteclient .search-noresult')
                solde_span = page.locator("#toolbar #btns-action .item-bloc-stat span:has-text('Factures') + strong .solde-price")
                nom_client = page.locator('.synthese-client-rowtop section .card-block > div > div:nth-child(1) > div.p-l-10 > div:nth-child(1) > div > label').inner_text()
                num_client = page.locator('.synthese-title h2 strong').inner_text()
                adresse_client = page.locator('#contrat-info-gen wrp-address > span').inner_text()
                agence = page.locator('#contrat-info-gen > div:nth-child(2) > div:nth-child(3) > label.form-control-label:nth-child(2)').inner_text().strip()
                agence = re.sub(r'\s*-\s*', '-', re.sub(r'\s+', ' ', agence).strip())
                if empty_solde.is_visible() and not solde_span.is_visible():

                    set_info_msgs_callback('info', f"le contrat {contrat} n'a Aucune Facture")
                    fetched_data = {}
                    num_fact_handle="Aucune Facture pour ce contrat"
                    montant_handle =  0
                    echeance_handle = "-"
                    pns_handle = "-"
                    statut = "-"
                    echeance_status = "-"

                    fetched_data['Contrat'] = contrat
                    fetched_data['N° Client'] = num_client
                    fetched_data['Nom Client'] = nom_client
                    fetched_data['Adresse Client'] = adresse_client
                    fetched_data['Agence'] = agence
                    fetched_data['Facture'] = num_fact_handle
                    fetched_data['Montant'] = montant_handle
                    fetched_data['Echéance'] = echeance_handle
                    fetched_data['Montant Exigible'] = pns_handle
                    fetched_data['Statut de Règlement'] = statut
                    fetched_data['Statut Echéance'] = echeance_status
                    fetched_data_list.append(fetched_data)
                else:
                    solde_val = solde_span.text_content()
                    solde_val = clean_price_spans(solde_val)
                    invoice_blocks = page.locator('#blocNC1 .hpanel')
                    solde_span_color = solde_span.evaluate('el=> window.getComputedStyle(el).color')
                    solde_type_text = "Accompte de -" if solde_span_color == 'rgb(120, 163, 0)' else "Montant Impayé de "
                    set_info_msgs_callback('info', f"le contrat : {contrat} a un {solde_type_text}{solde_val} DHS")

                    for j in range(invoice_blocks.count()): # Iterate through invoice blocks
                        invoice_block = invoice_blocks.nth(j)
                        fetched_data = {}
                        #Num facture
                        num_fact_handle = invoice_block.locator('.panel-body .text-muted a')
                        num_fact_handle =  num_fact_handle.inner_text() if num_fact_handle.count() > 0 else None
                        print(f"num_fact_handle : {num_fact_handle}")
                        #montant
                        montant_handle = invoice_block.locator('.panel-footer .row > div:first-child >  .item-bloc-stat strong')
                        montant_handle = clean_price_spans(montant_handle.inner_text()) if montant_handle.count() > 0 else None
                        print(f"montant_handle : {montant_handle}")
                        #echeance
                        echeance_handle = invoice_block.locator('.panel-footer .row > div:nth-child(4) > .item-bloc-stat strong')
                        echeance_handle = echeance_handle.inner_text() if echeance_handle.count() > 0 else None
                        if echeance_handle !="-" and echeance_handle != None:
                            echeance_handle = datetime.strptime(echeance_handle, "%d/%m/%Y")
                            echeance_status = "" if echeance_handle > datetime.today() else "E"
                            echeance_handle = echeance_handle.strftime("%d-%m-%Y")
                        else:
                            echeance_handle = "-"
                            echeance_status = ""
                        print(f"echeance_handle : {echeance_handle}")
                        print(f"echeance_status : {echeance_status}")
                        #pns
                        pns_handle = invoice_block.locator('.panel-body .solde-price')
                        if pns_handle.count() > 0:
                            is_negative = pns_handle.evaluate("el=>el.classList.contains('negative')")
                            pns_handle = clean_price_spans(pns_handle.inner_text())
                            statut ="Non Réglé" if is_negative else "Réglé"
                        else:
                            pns_handle = None # Ensure pns_handle is initialized
                            statut = "" # Ensure statut is initialized
                        print(f"pns_handle : {pns_handle}")
                        print(f"statut : {statut}")
                        fetched_data['Contrat'] = contrat
                        fetched_data['N° Client'] = num_client
                        fetched_data['Nom Client'] = nom_client
                        fetched_data['Adresse Client'] = adresse_client
                        fetched_data['Agence'] = agence
                        fetched_data ['Facture'] = num_fact_handle
                        fetched_data ['Montant'] = montant_handle
                        fetched_data ['Echéance'] = echeance_handle
                        fetched_data ['Montant Exigible'] = pns_handle
                        fetched_data ['Statut de Règlement'] = statut
                        fetched_data ['Statut Echéance'] = echeance_status
                        fetched_data_list.append(fetched_data)

            except TargetClosedError as e:
                set_info_msgs_callback('alert', f'Navigateur Fermé, avant de traiter le contrat {contrat} : \n Erreur : {e}')

            except TimeoutError as e:
                set_info_msgs_callback('alert', f'Délai d\'attente Dépassé pour le contrat {contrat} : \n Erreur : {e}')
                continue

            except Error as e:
                set_info_msgs_callback('alert', f'Erreur d\'extraction Serveur pour le contrat {contrat} : \n Erreur : {e}')
                continue

            except (AttributeError, TypeError) as e:
                set_info_msgs_callback('alert', f'Erreur de traitement de données pour le contrat {contrat} : \n Erreur : {e}')
                continue

            except ValueError as e:
                set_info_msgs_callback('alert', f'Erreur de conversion de données pour le contrat {contrat} : \n Erreur : {e}')
                continue

            except Exception as e:
                set_info_msgs_callback('alert', f'Erreur Inconnue pour le contrat {contrat} : \n Erreur : {e}')
                continue

            progress_percent = f'{((i + 1) / total_contrats) * 100:.0f}%'
            progress_text = f'{i + 1}/{total_contrats}'
            progress_callback(progress_text, progress_percent)

        # Clean up
        if browser: browser.close()
        if p: p.stop()
        set_info_msgs_callback('info', f"Données de {total_contrats} contrats extraites")
        print(f"Finally : {fetched_data_list}")
    return fetched_data_list

def execute(set_info_msgs_callback, extracted_data, check_value, progress_callback,app_instance):
    result_data = None
    if isinstance(extracted_data, list) and len(extracted_data) > 0:
        if all(isinstance(item, tuple) and len(item) == 2 for item in extracted_data):
            result_data = process_factures(set_info_msgs_callback, extracted_data, progress_callback, check_value)
            if result_data is not None:
                app_instance.print_button.configure(state="normal")
                set_info_msgs_callback('success', 'Données Factures Extraites à Partir de WAT.ERP')
                set_wf_state(app_instance, 4)
            else:
                set_info_msgs_callback('alert', 'Echec d\'extraction à Partir de WAT.ERP')

        # Check if it's a simple array of strings
        elif all(isinstance(item, str) for item in extracted_data):
            result_data = process_contrats(set_info_msgs_callback, extracted_data, progress_callback)
            if result_data is not None:
                app_instance.export_button.configure(state="normal")
                set_info_msgs_callback('success', 'Données Contrats Extraites à Partir de WAT.ERP')
                set_wf_state(app_instance, 4)
            else:
                set_info_msgs_callback('alert', 'Echec d\'extraction à Partir de WAT.ERP')

        else:
            set_info_msgs_callback('alert', 'Format de Données Non Reconnue')
            result_data = None

    return result_data


## 4 - EXPORT RESULTS TO EXCEL ##
# export_pns_to_execl needs set_info_msgs_callback
def export_pns_to_execl(set_info_msgs_callback, fetch_data_list):
    # Check if data exists and is valid
    if not fetch_data_list:
        set_info_msgs_callback('alert', 'Pas de données à Exporter, Merci d\'éxécuter le Processus d\'Extraction des Données WAT.ERP')
        return None

    # Ensure it's a list of dictionaries
    if not isinstance(fetch_data_list, list):
        set_info_msgs_callback('alert', f"Error: Expected list, got {type(fetch_data_list)}")
        return None
    print(f"Nombre d'enregistrements à exporter: {len(fetch_data_list)}")
    df = pd.DataFrame(fetch_data_list)
    print(fetch_data_list)


    for column in df.columns:
        try:
            converted = pd.to_numeric(df[column])
            df[column] = converted
        except (ValueError, TypeError):pass # Keep original column if conversion fails

    #File Names Format(Timestamps to prevent dduplicate files
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"PNS_Contrats_{timestamp}.xlsx"
    try:
        df.to_excel(filename, index=False, engine='openpyxl')
        wb = load_workbook(filename)
        ws = wb.active
        #Styling (Add Title and a blank row

        title_str = datetime.now().strftime('Situation de Paiement Au %d/%m/%Y')

        ws.insert_rows(1, amount=2) #One for Title and One blank

        #Style Title Row(Merge, color and Bold Font)
        ws.merge_cells (start_row = 1, start_column = 1, end_row = 1, end_column = ws.max_column)
        title_cell  = ws.cell(row = 1, column = 1)
        title_cell.value = title_str
        title_cell.font = Font(bold = True, size=13)
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')

        #Style Header Row (Color, Bold Font)
        header_fill = PatternFill(start_color= 'CCCCCC', end_color='CCCCCC', fill_type='solid')
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        for col in ws.iter_cols(min_row=3, max_row=3):
            for cell in col:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border



        montant_col_index = None
        exigible_col_index = None
        statut_col_index = None
        for idx, cell in enumerate(ws[3], start = 1):
            value = str(cell.value).strip().lower()
            if value == "montant":montant_col_index = idx
            if value == "montant exigible":exigible_col_index = idx
            if value == "statut de règlement":statut_col_index = idx

        merge_text = 'Aucune Facture pour ce contrat'

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ws.max_column):
            ws.row_dimensions[row[0].row].height = 18
            statut_value = row[statut_col_index-1].value if statut_col_index else None
            for idx, cell in enumerate(row, start = 1):
                cell.alignment = center_align
                cell.border = thin_border
                if idx == montant_col_index and isinstance(cell.value, (int, float)) or idx == exigible_col_index and isinstance(cell.value, (int, float)):

                    cell.number_format = '[$-40C]#,##0.00'

                if statut_col_index and idx == statut_col_index:
                    if str(statut_value).strip().lower() == "réglé": cell.font = Font(color = "008000")
                    elif str(statut_value).strip().lower() == "non réglé": cell.font = Font(color = "FF0000")

            for cell in row:
                if cell.value and str(cell.value).strip() == merge_text:
                    row_index = cell.row
                    start_col = cell.column
                    end_col = ws.max_column
                    ws.merge_cells(start_row = row_index, end_row = row_index, start_column = start_col, end_column = end_col)
                    merged_cell = ws.cell(row = row_index, column = start_col)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    break


        for column_cells in ws.columns:
            max_length = 0

            #Get the column letter
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            else:continue

            #Adjust the cols width
            for cell in column_cells:
                try:
                    if cell.value :max_length = max(max_length, len(str(cell.value)))
                except: pass

            adjusted_width = max_length
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filename)
        set_info_msgs_callback('success', f"\nDonnées Exportées avec Succès : {filename}")
        time.sleep(3)
        set_info_msgs_callback('info', f'Total d\'enregistrements exportés:"{len(fetch_data_list)}"')

        return filename

    except Exception as e:

        set_info_msgs_callback('alert', f"Erreur d'export Excel: {e}")
        time.sleep(3)
        # Convert to CSV if Excel fails

        #CSV File Name
        csv_filename = f"Données PNS _ {timestamp}.csv"
        df.to_csv(csv_filename, index=False)

        set_info_msgs_callback('info', f"Exported to CSV instead: {csv_filename}")
        time.sleep(3)
        return csv_filename


def print_test(app_instance, start_time, end_time, download_dir): # No change needed here if it doesn't call set_info_msgs
    fallback_dir = os.path.expanduser("~\\Downloads")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, executable_path=CHROME_PATH)
        context = browser.new_context(storage_state="state.json",
                                      ignore_https_errors=True, accept_downloads= True)
        context.set_default_navigation_timeout(60000)
        page = context.new_page()
        page.goto("https://waterp-cas.srm-cas.local")
        page.wait_for_selector("#editions ul li")
        page.wait_for_load_state("domcontentloaded")

        page.wait_for_timeout(5000)
        editions = page.locator("#editions ul li")

        #before clicking here we MUST check if editions is > 3 elements -_-
        #if (editions locator count > 3), do the following
        page.locator("#editions .dashmod--more--more").click()
        counter = 0

        while counter <10:
                icon_locator = editions.nth(counter).locator(".dashmod__editions--link i:nth-child(1)")
                has_fa_check = icon_locator.evaluate("el => el.classList.contains('fa-check')")
                if has_fa_check:
                    with page.expect_download() as download_info:
                        editions.nth(counter).locator(".dashmod__editions--link").click()
                    download = download_info.value
                    # download.save_as(f"C:\\Users\\star1\\Downloads\\{download.suggested_filename}")  # Removed redundant download var

                    try:
                        save_path = os.path.join(download_dir, download.suggested_filename)
                        download.save_as(save_path)
                    except Exception as e:
                        app_instance.set_info_msgs('alert', f"Erreur lors de l'enregistrement dans {download_dir}, enregistrement dans le dossier Téléchargements. Détail : {e}")
                        fallback_path = os.path.join(fallback_dir, download.suggested_filename)
                        download.save_as(fallback_path)

                counter = counter+1



        browser.close()
